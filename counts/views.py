import datetime
from collections import defaultdict
from io import BytesIO
from tempfile import NamedTemporaryFile
import os
from pathlib import Path
from django.http import HttpResponse
from django.shortcuts import render
from counts.models import Warehouse
from routes.logic import box
from routes.logic import adapter
from routes.logic import download_deliveries as dd
from .models import Invoice
from .forms import StartDateForm, EndDateForm
from .forms import UploadFileForm
import xlrd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def invoice(request):
    context = {
        "file_form": UploadFileForm(),
        "start_date_form": StartDateForm(),
        "end_date_form": EndDateForm(),
    }
    return render(request, "counts/invoice.html", context=context)


def replace_empty(val):
    return 0 if val == '' else val


def cols_to_values(row):
    return row[0].value, row[6].value


def extract_warehouse_tag(tup):
    return tup[0].split()[1]


def key_group(dictionaries, keys):
    if type(keys) != tuple:
        keys = (keys,)
    result = defaultdict(lambda: [])
    
    for dictionary in dictionaries:
        key = tuple(dictionary.get(key, 'missing') for key in keys)
        
        result[key].append(dictionary)
        
    return dict(result)


def is_countable(delivery):
    return ((delivery.get('delivery_status') == 'Completed') |
            (delivery.get('delivery_status') == 'Future'))


def filter_for_completed(deliveries):
    return [delivery for delivery in deliveries if is_countable(delivery)]


def get_estimate(start_date, end_date):
    download = dd.collect_time_span(start_date, end_date)
    cleaned = adapter.clean_unrouted(download)

    groups = key_group(cleaned, ('delivery_date', 'deliverytime'))

    translator = adapter.Translator()

    result = []
    for group, deliveries in groups.items():
        date, time_window = group
        countable = filter_for_completed(deliveries)
        print(len(deliveries) == len(countable))
        if whs := Warehouse.objects.filter(date=date, time_window=time_window):
            warehouse = adapter.build_warehouse_from_db(whs.first())
        else:
            warehouse = adapter.build_basic_warehouse()

        orders = [adapter.build_box_order(stop)(translator) for stop in countable]

        temp = box.sum_prototypes(
                [
                    box.to_prototype(box.build_box_from_order(order)(warehouse))
                    for order in orders
                ]
            )
        result = box.add_prototypes(result, temp)

    return result


def post_invoice(request):
    file = request.FILES.get('file')
    post = dict(request.POST.lists())
    start_date, end_date = tuple(
        datetime.datetime(int(year), int(month), int(day)) for year, month, day in zip(
            post["date_year"], 
            post["date_month"], 
            post["date_day"], 
        )
    )
    book = xlrd.open_workbook(file_contents=file.read())

    sh = book.sheet_by_index(0)

    path = Path(
        os.path.dirname(__file__),
        'static',
        'counts',
        'CountTemplate.xlsx'
    )

    workbook = load_workbook(path)

    next_warehouse = "empty warehouse"

    result = defaultdict(list)
    for rx in range(sh.nrows):
        tup = cols_to_values(sh.row(rx))
        if (tup[0].startswith('On Hand') | tup[0].startswith('Report')):
            continue
        if (tup[0].startswith('Warehouse')):
            next_warehouse = extract_warehouse_tag(tup)
            continue
        result[next_warehouse].append((tup[0], int(replace_empty(tup[1]))))

    for warehouse, items in result.items():
        worksheet = workbook.create_sheet(title=warehouse)
        for i, item in enumerate(items, 1):
            for j, col in enumerate(item, 1):
                letter = get_column_letter(j)
                worksheet[f"{letter}{i}"] = col
    
    estimate = get_estimate(start_date, end_date)

    worksheet = workbook.create_sheet(title='Estimated')
    for i, item in enumerate(estimate, 1):
        for j, col in enumerate(item, 1):
            letter = get_column_letter(j)
            worksheet[f"{letter}{i}"] = col

    worksheet = workbook['Summary']
    # repair broken linkes
    for i in range(68):
        worksheet[f"D{3 + i}"] = f"=IFERROR(VLOOKUP(A{i + 3},DET!$A$1:$C$61,2,FALSE), 0)"
        worksheet[f"L{3 + i}"] =  f"=IFERROR(VLOOKUP(A{i + 3},MG!$A$1:$C$61,2,FALSE), 0)"
        worksheet[f"M{3 + i}"] =  f"=IFERROR(VLOOKUP(A{i + 3},MGT1!$A$1:$C$61,2,FALSE), 0)" 
        worksheet[f"I{3 + i}"] =  f"=IFERROR(VLOOKUP(A{i + 3},Estimated!$A$1:$C$61,2,FALSE), 0)" 
        worksheet[f"G{3 + i}"] =  f"=E{i + 3}*C{i + 3}+F{i + 3}" 
        worksheet[f"J{3 + i}"] =  f"=D{i + 3}-G{i + 3}" 

    f = NamedTemporaryFile(delete=False)
    workbook.save(f.name)
    output = BytesIO(f.read())
    f.close()
    os.unlink(f.name)

    Invoice.objects.create(start_date=start_date, end_date=end_date)

    response = HttpResponse(output)
    response[
        "Content-Disposition"
    ] = f'attachment; filename="CountWorksheet{start_date.strftime("%Y-%m-%d")}through{end_date.strftime("%Y-%m-%d")}.xlsx"'
    
    return response
